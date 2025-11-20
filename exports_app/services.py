"""
Servicios para exportar datos a Excel usando plantillas.
"""
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import json
from django.http import HttpResponse
from pathlib import Path
from typing import Dict, Any
from .data_collectors import FichaAPIDataCollector, ProyectoAPIDataCollector


def _normalize_coord(coord: str) -> str:
    """
    Accept a single cell (e.g. "C5") or a range (e.g. "C5:D5").
    Return the top-left coordinate (e.g. "C5").
    """
    if not coord:
        return coord
    coord = str(coord).strip()
    if ":" in coord:
        start, _ = coord.split(":", 1)
        return start.strip()
    return coord


def _set_value_safe(ws, coord: str, value):
    """Set value in a cell, handling merged cells correctly."""
    coord = _normalize_coord(coord)
    if not coord:
        return
        
    try:
        cell = ws[coord]
        # If targeting a merged cell that is not the top-left anchor, redirect
        if isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if coord in rng:
                    anchor = ws.cell(row=rng.min_row, column=rng.min_col)
                    anchor.value = value
                    return
            # If not found in ranges (edge case), just skip silently
            return
        cell.value = value
    except Exception as e:
        # Log error but don't fail the entire export
        print(f"Error setting value at {coord}: {e}")


def export_ficha_api(subject) -> HttpResponse:
    """
    Exporta una Ficha API completa para una asignatura.
    
    Args:
        subject: Objeto Subject con todos sus datos relacionados
        
    Returns:
        HttpResponse con el archivo Excel
        
    NOTA: Este sistema SIEMPRE genera un Excel válido, incluso si faltan datos.
    Los campos faltantes se rellenan con cadenas vacías.
    """
    # 1. Obtener ruta de plantilla y mapeo
    template_path = get_template_path('ficha-api')
    mapping_path = get_mapping_path('ficha-api')
    
    # 2. Cargar plantilla Excel
    wb = load_workbook(filename=template_path)
    ws = wb.active
    
    # 3. Recolectar datos de la base de datos
    collector = FichaAPIDataCollector(subject)
    data = collector.collect_all()
    
    # 4. Cargar mapeo JSON
    with open(mapping_path, 'r', encoding='utf-8') as f:
        cell_mapping = json.load(f)
    
    # 5. Llenar celdas según el mapeo
    for field_key, cell_coord in cell_mapping.items():
        value = data.get(field_key, '')
        _set_value_safe(ws, cell_coord, value)
    
    # 6. Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Ficha_API_{subject.code}_{subject.period_season}{subject.period_year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 7. Agregar información sobre datos faltantes en headers
    missing_data = collector.get_missing_data_report()
    if missing_data:
        response['X-Export-Status'] = 'partial'
        response['X-Missing-Data-Count'] = str(len(missing_data))
        # Opcional: agregar lista de datos faltantes (limitado por tamaño de header)
        # response['X-Missing-Data'] = '; '.join(missing_data[:3])
    else:
        response['X-Export-Status'] = 'complete'
    
    # 8. Guardar y retornar
    wb.save(response)
    return response


def get_template_path(template_key: str) -> str:
    """Obtiene la ruta de la plantilla Excel."""
    base = Path(__file__).resolve().parent / "templates" / "excel"
    templates = {
        "ficha-api": "ficha_api.xlsx",
        "proyecto-api": "proyecto_api.xlsx",
    }
    if template_key not in templates:
        raise ValueError(f"Template '{template_key}' no encontrado")
    return str(base / templates[template_key])


def get_mapping_path(template_key: str) -> str:
    """Obtiene la ruta del archivo JSON de mapeo."""
    base = Path(__file__).resolve().parent / "templates" / "excel"
    mappings = {
        "ficha-api": "ficha_api_celdas_de_respuestas_mapeadas.json",
        "proyecto-api": "proyecto_api_celdas_de_respuestas_mapeadas.json",
    }
    if template_key not in mappings:
        raise ValueError(f"Mapping '{template_key}' no encontrado")
    return str(base / mappings[template_key])


def export_proyecto_api(subject, problem_statement) -> HttpResponse:
    """
    Exporta una Ficha Proyecto API completa para un proyecto específico de una asignatura.
    
    Args:
        subject: Objeto Subject con todos sus datos relacionados
        problem_statement: Objeto ProblemStatement específico a exportar
        
    Returns:
        HttpResponse con el archivo Excel
        
    NOTA: Este sistema SIEMPRE genera un Excel válido, incluso si faltan datos.
    Los campos faltantes se rellenan con cadenas vacías.
    """
    # 1. Obtener ruta de plantilla y mapeo
    template_path = get_template_path('proyecto-api')
    mapping_path = get_mapping_path('proyecto-api')
    
    # 2. Cargar plantilla Excel
    wb = load_workbook(filename=template_path)
    ws = wb.active
    
    # 3. Recolectar datos de la base de datos
    collector = ProyectoAPIDataCollector(subject, problem_statement)
    data = collector.collect_all()
    
    # 4. Cargar mapeo JSON
    with open(mapping_path, 'r', encoding='utf-8') as f:
        cell_mapping = json.load(f)
    
    # 5. Llenar celdas según el mapeo
    for field_key, cell_coord in cell_mapping.items():
        value = data.get(field_key, '')
        _set_value_safe(ws, cell_coord, value)
    
    # 6. Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Incluir información de la empresa en el nombre del archivo
    company_name = problem_statement.company.name.replace(' ', '_').replace('/', '_')[:30]
    filename = f"Proyecto_API_{subject.code}_{subject.section}_{company_name}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 7. Agregar información sobre datos faltantes en headers
    missing_data = collector.get_missing_data_report()
    if missing_data:
        response['X-Export-Status'] = 'partial'
        response['X-Missing-Data-Count'] = str(len(missing_data))
    else:
        response['X-Export-Status'] = 'complete'
    
    # 8. Guardar y retornar
    wb.save(response)
    return response
